#!/usr/bin/env python3
"""
Job Tracker - Simplified Version for Python 3.12
A Flask web application for tracking job applications locally with SQLite database.
This version uses openpyxl directly instead of pandas to avoid compatibility issues.
"""

import os
import sqlite3
import threading
import time
import webbrowser
from datetime import datetime, timedelta
from urllib.parse import urlparse
import re
import csv
import io

import pyperclip
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file, make_response
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook

# Optional web scraping imports
try:
    import requests
    from bs4 import BeautifulSoup
    from urllib.robotparser import RobotFileParser
    SCRAPING_AVAILABLE = True
    print("Web scraping capabilities enabled")
except ImportError:
    SCRAPING_AVAILABLE = False
    print("Web scraping dependencies not installed. Install with: pip install requests beautifulsoup4 lxml")

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'

@app.template_filter('fmt_date')
def fmt_date(value):
    if not value:
        return ''
    date_str = str(value)[:10]
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').strftime('%d %b %Y')
    except ValueError:
        return value

# Configuration
DATABASE_PATH = os.environ.get("DATABASE_PATH", "data/jobtracker.sqlite3")
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Ensure directories exist
db_dir = os.path.dirname(DATABASE_PATH)
if db_dir:
    os.makedirs(db_dir, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Global variables for clipboard monitoring
clipboard_thread = None
stop_clipboard_monitoring = False
last_clipboard_content = ""
pending_draft = None

class JobTracker:
    def __init__(self, db_path=DATABASE_PATH):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize the SQLite database with required tables."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS applications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            role TEXT NOT NULL,
            job_id TEXT,
            date_applied DATE DEFAULT CURRENT_DATE,
            status TEXT DEFAULT 'Applied',
            url TEXT UNIQUE,
            platform TEXT,
            location TEXT,
            salary TEXT,
            notes TEXT,
            tags TEXT,
            next_action TEXT,
            follow_up_on DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        
        # Create index for faster URL lookups
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_url ON applications(url)')
        
        conn.commit()
        conn.close()

    def get_connection(self):
        """Get database connection."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn
    
    def add_application(self, data):
        """Add a new job application."""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
            INSERT INTO applications 
            (company, role, job_id, date_applied, status, url, platform, 
             location, salary, notes, tags, next_action, follow_up_on)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data.get('company', ''),
                data.get('role', ''),
                data.get('job_id', ''),
                data.get('date_applied', datetime.now().strftime('%Y-%m-%d')),
                data.get('status', 'Applied'),
                data.get('url', ''),
                data.get('platform', ''),
                data.get('location', ''),
                data.get('salary', ''),
                data.get('notes', ''),
                data.get('tags', ''),
                data.get('next_action', ''),
                data.get('follow_up_on', None) if data.get('follow_up_on') else None
            ))
            conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            return None  # URL already exists
        finally:
            conn.close()
    
    def get_all_applications(self):
        """Get all job applications."""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
        SELECT * FROM applications 
        ORDER BY date_applied DESC, created_at DESC
        ''')
        applications = cursor.fetchall()
        conn.close()
        return [dict(app) for app in applications]
    
    def get_application(self, app_id):
        """Get a specific application by ID."""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM applications WHERE id = ?', (app_id,))
        app = cursor.fetchone()
        conn.close()
        return dict(app) if app else None
    
    def update_application(self, app_id, data):
        """Update an existing application."""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Build dynamic update query
        fields = []
        values = []
        for key, value in data.items():
            if key != 'id':
                fields.append(f"{key} = ?")
                values.append(value)
        
        if fields:
            fields.append("updated_at = ?")
            values.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            values.append(app_id)
            
            query = f"UPDATE applications SET {', '.join(fields)} WHERE id = ?"
            cursor.execute(query, values)
            conn.commit()
        
        conn.close()
    
    def delete_application(self, app_id):
        """Delete an application."""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM applications WHERE id = ?', (app_id,))
        conn.commit()
        conn.close()
    
    def export_to_excel(self, filename='applications.xlsx'):
        """Export all applications to Excel using openpyxl."""
        applications = self.get_all_applications()
        if not applications:
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Job Applications"
            
            # Define headers
            headers = [
                'company', 'role', 'job_id', 'date_applied', 'status', 'url',
                'platform', 'location', 'salary', 'notes', 'tags', 'next_action', 'follow_up_on'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            
            # Write data
            for row, app in enumerate(applications, 2):
                for col, header in enumerate(headers, 1):
                    value = app.get(header, '')
                    ws.cell(row=row, column=col, value=value)
            
            wb.save(filename)
            return True
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def import_from_excel(self, filename):
        """Import applications from Excel file using openpyxl."""
        try:
            wb = load_workbook(filename)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    # Normalize header names
                    header = str(cell.value).lower().replace(' ', '_')
                    headers.append(header)
            
            imported_count = 0
            
            # Process each row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                    
                data = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        if value is not None:
                            data[headers[i]] = str(value)
                        else:
                            data[headers[i]] = ''
                
                # Ensure required fields
                if not data.get('company'):
                    data['company'] = 'Unknown Company'
                if not data.get('role'):
                    data['role'] = 'Unknown Role'
                
                # Try to add the application
                if self.add_application(data):
                    imported_count += 1
            
            return imported_count
        except Exception as e:
            print(f"Error importing Excel file: {e}")
            return 0

@app.route('/check_draft')
def check_draft():
    """Check for pending draft from clipboard monitoring."""
    global pending_draft
    
    if pending_draft:
        draft = pending_draft
        pending_draft = None  # Clear the pending draft
        return jsonify({'has_draft': True, 'draft': draft})
    
    return jsonify({'has_draft': False})

# Save draft route should be before main execution
@app.route('/save_draft', methods=['POST'])
def save_draft():
    """Save a draft application from clipboard monitoring."""
    data = request.get_json()
    
    if data:
        app_id = job_tracker.add_application(data)
        if app_id:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'URL already exists'})
    
    return jsonify({'success': False, 'error': 'No data provided'})


def clipboard_monitor():
    """Monitor clipboard for job URLs and create draft entries."""
    global last_clipboard_content, pending_draft, stop_clipboard_monitoring
    
    while not stop_clipboard_monitoring:
        try:
            current_content = pyperclip.paste()
            
            # Check if clipboard content changed and looks like a job URL
            if (current_content != last_clipboard_content and 
                current_content.startswith(('http://', 'https://')) and
                is_job_url(current_content)):
                
                # Extract job info from URL
                job_info = extract_job_info_from_url(current_content)
                
                # Create draft entry
                draft = {
                    'url': current_content,
                    'platform': job_info['platform'],
                    'company': job_info['company'],
                    'job_id': job_info['job_id'],
                    'role': job_info['role'],  # Now potentially filled by scraping
                    'location': job_info['location'],  # Now potentially filled by scraping
                    'salary': job_info['salary'],  # Now potentially filled by scraping
                    'date_applied': datetime.now().strftime('%Y-%m-%d'),
                    'status': 'Applied'
                }
                
                pending_draft = draft
                print(f"Draft created for: {current_content}")
            
            last_clipboard_content = current_content
        except Exception as e:
            print(f"Clipboard monitoring error: {e}")
        
        time.sleep(1)  # Check every second


# Initialize job tracker
job_tracker = JobTracker()

def is_interview_status(status):
    if not status:
        return False
    status_lower = status.lower()
    keywords = ['interview', 'technical', 'final', 'assessment']
    return any(keyword in status_lower for keyword in keywords)

def extract_job_info_from_url(url):
    """Extract job information from URL patterns and optionally web scraping."""
    parsed_url = urlparse(url)
    domain = parsed_url.netloc.lower()
    
    job_info = {
        'platform': '',
        'company': '',
        'job_id': '',
        'role': '',
        'location': '',
        'salary': ''
    }
    
    # LinkedIn
    if 'linkedin.com' in domain:
        job_info['platform'] = 'LinkedIn'
        # Extract job ID from LinkedIn URL
        job_id_match = re.search(r'/jobs/view/(\d+)', url)
        if job_id_match:
            job_info['job_id'] = job_id_match.group(1)
    
    # Greenhouse
    elif 'greenhouse.io' in domain or 'boards.greenhouse.io' in domain:
        job_info['platform'] = 'Greenhouse'
        # Extract company from subdomain
        company_match = re.search(r'https?://([^.]+)\.greenhouse\.io', url)
        if company_match:
            job_info['company'] = company_match.group(1).replace('-', ' ').title()
        # Extract job ID
        job_id_match = re.search(r'/jobs/(\d+)', url)
        if job_id_match:
            job_info['job_id'] = job_id_match.group(1)
    
    # Lever
    elif 'lever.co' in domain:
        job_info['platform'] = 'Lever'
        # Extract company from subdomain
        company_match = re.search(r'https?://jobs\.lever\.co/([^/]+)', url)
        if company_match:
            job_info['company'] = company_match.group(1).replace('-', ' ').title()
    
    # Workday
    elif 'workday.com' in domain or 'myworkdayjobs.com' in domain:
        job_info['platform'] = 'Workday'
        # Try to extract company name from URL path
        company_match = re.search(r'/([^/]+)/job/', url)
        if company_match:
            job_info['company'] = company_match.group(1).replace('-', ' ').title()
    
    # Indeed
    elif 'indeed.com' in domain:
        job_info['platform'] = 'Indeed'
    
    # AngelList/Wellfound
    elif 'angel.co' in domain or 'wellfound.com' in domain:
        job_info['platform'] = 'AngelList'
    
    # Glassdoor
    elif 'glassdoor.com' in domain:
        job_info['platform'] = 'Glassdoor'
    
    # Company career pages
    elif any(keyword in domain for keyword in ['careers', 'jobs']):
        job_info['platform'] = 'Company Website'
        # Try to extract company from domain
        domain_parts = domain.replace('careers.', '').replace('jobs.', '').split('.')
        if len(domain_parts) >= 2:
            job_info['company'] = domain_parts[-2].title()
    
    # Generic job boards
    elif any(keyword in domain for keyword in ['hiring', 'employment', 'position']):
        job_info['platform'] = 'Job Board'
    
    # Attempt web scraping if available and appropriate
    if SCRAPING_AVAILABLE and should_attempt_scraping(url):
        try:
            scraped_info = scrape_job_details(url)
            # Merge scraped info, preferring existing URL-parsed data
            for key, value in scraped_info.items():
                if value and not job_info.get(key):
                    job_info[key] = value
        except Exception as e:
            print(f"Scraping failed for {url}: {e}")
    
    return job_info

def should_attempt_scraping(url):
    """Determine if we should attempt scraping for this URL."""
    if not SCRAPING_AVAILABLE:
        return False
    
    # Only scrape from sites likely to allow it
    safe_patterns = [
        r'.*careers\..*',
        r'.*jobs\..*',
        r'.*\.careers\..*',
        # Add more patterns for sites you know allow scraping
    ]
    
    return any(re.match(pattern, url, re.IGNORECASE) for pattern in safe_patterns)

def scrape_job_details(url):
    """Scrape job details from a URL using ethical practices."""
    if not SCRAPING_AVAILABLE:
        return {}
    
    try:
        # Check robots.txt first
        parsed_url = urlparse(url)
        base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
        robots_url = f"{base_url}/robots.txt"
        
        try:
            from urllib.robotparser import RobotFileParser
            rp = RobotFileParser()
            rp.set_url(robots_url)
            rp.read()
            
            if not rp.can_fetch('*', url):
                print(f"Robots.txt disallows scraping {url}")
                return {}
        except:
            # If robots.txt check fails, proceed cautiously
            pass
        
        # Make request with respectful headers and timeout
        headers = {
            'User-Agent': 'Job-Tracker/1.0 (Educational Project)',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
        }
        
        # Add delay to be respectful
        time.sleep(2)
        
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        scraped_info = {
            'company': '',
            'role': '',
            'location': '',
            'salary': ''
        }
        
        # Try common selectors for job information
        # Job title
        title_selectors = [
            'h1.job-title', 'h1[data-job-title]', '.job-title',
            'h1.position-title', '.position-title', '.role-title',
            'h1:first-of-type', 'title'
        ]
        
        for selector in title_selectors:
            element = soup.select_one(selector)
            if element and element.get_text().strip():
                title_text = element.get_text().strip()
                # Clean up title
                title_text = re.sub(r'^(Job:|Position:|Role:)\s*', '', title_text, flags=re.IGNORECASE)
                title_text = re.sub(r'\s*-\s*.*$', '', title_text)
                scraped_info['role'] = title_text
                break

        # Company name (if not already extracted from URL)
        company_selectors = [
            '.company-name', '.employer-name', '[data-company]',
            '.company', '.employer'
        ]
        
        for selector in company_selectors:
            element = soup.select_one(selector)
            if element and element.get_text().strip():
                scraped_info['company'] = element.get_text().strip()
                break
        
        # Location
        location_selectors = [
            '.job-location', '.location', '[data-location]',
            '.job-location-text', '.position-location'
        ]
        
        for selector in location_selectors:
            element = soup.select_one(selector)
            if element and element.get_text().strip():
                scraped_info['location'] = element.get_text().strip()
                break
        
        # Salary (if available)
        salary_selectors = [
            '.salary', '.compensation', '[data-salary]',
            '.salary-range', '.pay-range'
        ]
        
        for selector in salary_selectors:
            element = soup.select_one(selector)
            if element and element.get_text().strip():
                scraped_info['salary'] = element.get_text().strip()
                break
        
        print(f"Successfully scraped job info from {url}")
        return scraped_info
        
    except Exception as e:
        print(f"Error scraping {url}: {e}")
        return {}

def is_job_url(url):
    """Check if URL looks like a job posting."""
    if not url or not isinstance(url, str):
        return False
    
    # Check for common job-related keywords in URL
    job_keywords = [
        'job', 'career', 'hiring', 'employment', 'position',
        'linkedin.com/jobs', 'greenhouse.io', 'lever.co',
        'workday.com', 'indeed.com', 'glassdoor.com',
        'angel.co', 'wellfound.com'
    ]
    
    url_lower = url.lower()
    return any(keyword in url_lower for keyword in job_keywords)

# Flask Routes

@app.route('/')
def index():
    """Main dashboard showing all applications."""
    applications = job_tracker.get_all_applications()
    
    # Calculate comprehensive stats
    stats = {
        'total': len(applications),
        'applied': len([a for a in applications if a['status'] == 'Applied']),
        'interview': len([a for a in applications if is_interview_status(a['status'])]),
        'offer': len([a for a in applications if a['status'] in ['Offer', 'Accepted']]),
        'rejected': len([a for a in applications if a['status'] in ['Rejected', 'No Response']])
    }
    
    return render_template('dashboard.html', applications=applications, stats=stats)

@app.route('/analytics')
def analytics():
    """Analytics page with visualizations."""
    applications = job_tracker.get_all_applications()
    
    # Custom range support
    def _parse_date(value):
        if not value:
            return None
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            return None

    today_date = datetime.now().date()
    default_start = today_date - timedelta(days=29)
    start_param = request.args.get('start')
    end_param = request.args.get('end')
    start_date = _parse_date(start_param) or default_start
    end_date = _parse_date(end_param) or today_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    if not applications:
        # Return empty analytics for no applications
        stats = {'total': 0, 'response_rate': 0, 'interview_rate': 0, 'success_rate': 0}
        flow_data = {'no_response': 0, 'rejected_early': 0, 'interviews': 0, 'no_offer': 0, 'offers': 0, 'declined': 0, 'accepted': 0}
        total_days = (end_date - start_date).days + 1
        timeline_dates = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(total_days)]
        timeline_counts = [0] * total_days
        return render_template('analytics.html', stats=stats, flow_data=flow_data, insights=[], follow_ups=[],
                             platform_names=[], platform_counts=[], timeline_dates=timeline_dates,
                             timeline_counts=timeline_counts,
                             start_date=start_date.strftime('%Y-%m-%d'),
                             end_date=end_date.strftime('%Y-%m-%d'))
    
    # Calculate comprehensive stats
    total = len(applications)
    applied = len([a for a in applications if a['status'] == 'Applied'])
    interviews = len([a for a in applications if is_interview_status(a['status'])])
    offers = len([a for a in applications if a['status'] in ['Offer', 'Accepted']])
    rejected = len([a for a in applications if a['status'] == 'Rejected'])
    no_response = len([a for a in applications if a['status'] == 'No Response'])
    
    # Calculate rates
    response_rate = round(((interviews + offers + rejected) / total * 100), 1) if total > 0 else 0
    interview_rate = round((interviews / total * 100), 1) if total > 0 else 0
    success_rate = round((offers / total * 100), 1) if total > 0 else 0
    
    stats = {
        'total': total,
        'applied': applied,
        'interview': interviews,
        'offer': offers,
        'rejected': rejected,
        'no_response': no_response,
        'response_rate': response_rate,
        'interview_rate': interview_rate,
        'success_rate': success_rate
    }
    
    # Sankey flow data
    flow_data = {
        'no_response': no_response,
        'rejected_early': rejected,
        'interviews': interviews,
        'no_offer': len([a for a in applications if is_interview_status(a['status']) and a['status'] not in ['Offer', 'Accepted']]),
        'offers': offers,
        'declined': len([a for a in applications if a['status'] == 'Declined']),
        'accepted': len([a for a in applications if a['status'] == 'Accepted'])
    }
    
    # Platform analysis
    platform_counts = {}
    for app in applications:
        platform = app.get('platform', 'Unknown')
        if platform:
            platform_counts[platform] = platform_counts.get(platform, 0) + 1
    
    platform_names = list(platform_counts.keys())
    platform_values = list(platform_counts.values())
    
    # Timeline analysis (applications per day - custom range)
    from collections import defaultdict
    timeline_data = defaultdict(int)
    
    for app in applications:
        if app.get('date_applied'):
            try:
                date_str = app['date_applied'][:10]  # YYYY-MM-DD
                app_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                if start_date <= app_date <= end_date:
                    timeline_data[date_str] += 1
            except:
                pass
    
    timeline_dates = []
    timeline_counts = []
    total_days = (end_date - start_date).days + 1
    for i in range(total_days):
        day = start_date + timedelta(days=i)
        day_str = day.strftime('%Y-%m-%d')
        timeline_dates.append(day_str)
        timeline_counts.append(timeline_data.get(day_str, 0))
    
    # Generate insights
    insights = generate_insights(applications, stats)
    
    # Follow-ups due
    follow_ups = []
    today = datetime.now().strftime('%Y-%m-%d')
    
    for app in applications:
        if app.get('follow_up_on') and app['follow_up_on'] <= today:
            follow_ups.append(app)
    
    follow_ups = follow_ups[:5]  # Limit to 5 most recent
    
    return render_template('analytics.html', 
                         stats=stats, 
                         flow_data=flow_data,
                         insights=insights,
                         follow_ups=follow_ups,
                         platform_names=platform_names,
                         platform_counts=platform_values,
                         timeline_dates=timeline_dates,
                         timeline_counts=timeline_counts,
                         start_date=start_date.strftime('%Y-%m-%d'),
                         end_date=end_date.strftime('%Y-%m-%d'))

def generate_insights(applications, stats):
    """Generate smart insights based on application data."""
    insights = []
    
    if stats['total'] == 0:
        return insights
    
    # Response rate insights
    if stats['response_rate'] < 20:
        insights.append({
            'type': 'warning',
            'icon': 'exclamation-triangle',
            'title': 'Low Response Rate',
            'message': f'Your response rate is {stats["response_rate"]}%. Consider improving your resume or targeting more suitable positions.'
        })
    elif stats['response_rate'] > 50:
        insights.append({
            'type': 'success',
            'icon': 'check-circle',
            'title': 'Great Response Rate!',
            'message': f'Your {stats["response_rate"]}% response rate is excellent. Keep up the good work!'
        })
    
    # Application volume insights
    if stats['total'] < 10:
        insights.append({
            'type': 'info',
            'icon': 'info-circle',
            'title': 'Consider Applying More',
            'message': 'Job searching is a numbers game. Consider increasing your application volume for better results.'
        })
    elif stats['total'] > 50:
        insights.append({
            'type': 'primary',
            'icon': 'trophy',
            'title': 'High Application Volume',
            'message': f'You\'ve submitted {stats["total"]} applications - great persistence!'
        })
    
    # Interview conversion insights
    if stats['interview'] > 0 and stats['offer'] == 0:
        insights.append({
            'type': 'warning',
            'icon': 'arrow-up',
            'title': 'Interview Skills Focus',
            'message': 'You\'re getting interviews but no offers. Consider practicing interview skills or researching companies better.'
        })
    
    # Platform performance
    platform_counts = {}
    for app in applications:
        platform = app.get('platform', 'Unknown')
        if platform:
            platform_counts[platform] = platform_counts.get(platform, 0) + 1
    
    if platform_counts:
        best_platform = max(platform_counts, key=platform_counts.get)
        insights.append({
            'type': 'info',
            'icon': 'graph-up',
            'title': 'Platform Performance',
            'message': f'{best_platform} is your most used platform with {platform_counts[best_platform]} applications.'
        })
    
    return insights[:4]  # Limit to 4 insights

@app.route('/dashboard')
def dashboard():
    """Alternative dashboard route."""
    return index()

@app.route('/add', methods=['GET', 'POST'])
def add_application():
    """Add a new job application."""
    if request.method == 'POST':
        data = request.form.to_dict()
        app_id = job_tracker.add_application(data)
        
        if app_id:
            flash('Application added successfully!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Error: Application with this URL already exists!', 'error')
    
    # Pass today's date to template
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('add_application.html', today=today)

@app.route('/edit/<int:app_id>', methods=['GET', 'POST'])
def edit_application(app_id):
    """Edit an existing application."""
    application = job_tracker.get_application(app_id)
    if not application:
        flash('Application not found!', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        data = request.form.to_dict()
        job_tracker.update_application(app_id, data)
        flash('Application updated successfully!', 'success')
        return redirect(url_for('index'))
    
    return render_template('edit_application.html', application=application)

@app.route('/delete/<int:app_id>', methods=['POST'])
def delete_application(app_id):
    """Delete an application."""
    job_tracker.delete_application(app_id)
    flash('Application deleted successfully!', 'success')
    return redirect(url_for('index'))

@app.route('/update_status', methods=['POST'])
def update_status():
    """Update application status via AJAX."""
    data = request.get_json()
    app_id = data.get('id')
    status = data.get('status')
    
    if app_id and status:
        job_tracker.update_application(app_id, {'status': status})
        return jsonify({'success': True})
    
    return jsonify({'success': False})

@app.route('/export')
def export_excel():
    """Export applications to Excel."""
    filename = f"applications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('data', filename)
    
    if job_tracker.export_to_excel(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    else:
        flash('No applications to export!', 'error')
        return redirect(url_for('index'))

@app.route('/import', methods=['GET', 'POST'])
def import_excel():
    """Import applications from Excel."""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected!', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected!', 'error')
            return redirect(request.url)
        
        if file and file.filename.lower().endswith(('.xlsx', '.xls')):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            imported_count = job_tracker.import_from_excel(filepath)
            
            # Clean up uploaded file
            os.remove(filepath)
            
            if imported_count > 0:
                flash(f'Successfully imported {imported_count} applications!', 'success')
            else:
                flash('No new applications were imported (duplicates skipped).', 'info')
            
            return redirect(url_for('index'))
        else:
            flash('Invalid file format. Please upload an Excel file (.xlsx or .xls).', 'error')
    
    return render_template('import.html')


# Move the main execution block here, at the very end
if __name__ == '__main__':
    # Start clipboard monitoring in background thread
    clipboard_thread = threading.Thread(target=clipboard_monitor, daemon=True)
    clipboard_thread.start()
    
    # Open browser automatically
    def open_browser():
        time.sleep(1)  # Wait for server to start
        webbrowser.open('http://127.0.0.1:5001')
    
    browser_thread = threading.Thread(target=open_browser, daemon=True)
    browser_thread.start()
    
    print("🚀 Job Tracker starting...")
    print("📋 Clipboard monitoring active - copy job URLs to auto-create drafts!")
    print("🌐 Opening browser at http://127.0.0.1:5001")
    
    try:
        port = int(os.environ.get("PORT", "5001"))
        app.run(debug=False, use_reloader=False, host="0.0.0.0", port=port)
    except KeyboardInterrupt:
        print("\n👋 Shutting down Job Tracker...")
        stop_clipboard_monitoring = True
